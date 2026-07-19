#!/usr/bin/env python3
"""Parse Microbiology department planners into the canonical task register.

Follows master-reminder-system/PARSING_SPEC.md exactly.
Department string = "Micro". Reference date 2026-07-15.
Occurrence window: 2026-01-01 .. 2027-07-31.
"""
import os
import re
import csv
from datetime import datetime, date

import xlrd
import openpyxl

ROOT = "/Users/swarali/Desktop/Enicar/ENincar 2025-26 schedule"
OUT_CSV = os.path.join(ROOT, "master-reminder-system/data/parsed/micro_register.csv")
OUT_EXC = os.path.join(ROOT, "master-reminder-system/data/parsed/micro-exceptions.md")

WIN_START = date(2026, 1, 1)
WIN_END = date(2027, 7, 31)

COLUMNS = [
    "task_id", "department", "planner_type", "task_name", "equipment_or_item_id",
    "activity_type", "frequency", "due_type", "due_date", "last_done_date",
    "done_date", "status", "responsible_email", "report_link", "remarks", "source_file",
]

rows = []          # list of dict
exceptions = []    # list of (source_file, sheet, rownum, cell, reason)

MONTHS = ["Apr", "May", "June", "July", "Aug", "Sept", "Oct", "Nov", "Dec", "Jan", "Feb", "Mar"]
ID_RE = re.compile(r"\(([A-Za-z]+/\d+)\)")

FREQ_MAP = {
    "yearly": "yearly",
    "quarterly": "quarterly",
    "half yearly": "half-yearly",
    "halfyearly": "half-yearly",
}


def norm_freq(raw):
    key = re.sub(r"\s+", " ", raw.strip().lower())
    key = key.replace("(outside agency)", "").strip()
    return FREQ_MAP.get(key.replace(" ", "") if key.replace(" ", "") in FREQ_MAP else key, key or "")


def in_window(d):
    return WIN_START <= d <= WIN_END


def xls_date(book, sheet, r, c):
    """Return a date if the cell is a date type, else None."""
    if sheet.cell_type(r, c) == xlrd.XL_CELL_DATE:
        try:
            return xlrd.xldate_as_datetime(sheet.cell_value(r, c), book.datemode).date()
        except Exception:
            return None
    return None


# ---------------------------------------------------------------------------
# 1. Calibration & Re-validation planners (legacy .xls matrix)
# ---------------------------------------------------------------------------
CAL_FILE = "Micro/Calibration & validation  Planner_26_27_Micro.xls"


def parse_matrix(sheet_name, planner_code, planner_type, activity_type):
    book = xlrd.open_workbook(os.path.join(ROOT, CAL_FILE))
    sheet = book.sheet_by_name(sheet_name)
    nnn = 0
    r = 2  # first data row (after title + header)
    while r < sheet.nrows:
        # A "Calib. On" row starts an instrument block; the next row is "Due On".
        param = str(sheet.cell_value(r, 3)).strip().lower()
        if not param.startswith("calib"):
            r += 1
            continue
        calib_row = r
        due_row = r + 1

        name = str(sheet.cell_value(calib_row, 1)).strip()
        freq_raw = str(sheet.cell_value(calib_row, 2)).strip()
        # ID: prefer the Due On row col1 (e.g. "(QCM/02011)"), else search name.
        id_src = str(sheet.cell_value(due_row, 1)) if due_row < sheet.nrows else ""
        m = ID_RE.search(id_src) or ID_RE.search(name)
        equip_id = m.group(1) if m else ""
        # Clean instrument name (strip trailing inline ID + whitespace)
        clean_name = ID_RE.sub("", name).strip()
        clean_name = re.sub(r"\s+", " ", clean_name)

        # outside-agency flag lives in the freq column of either row
        agency = ""
        for rr in (calib_row, due_row):
            for cc in (2, 3):
                if "outside agency" in str(sheet.cell_value(rr, cc)).lower():
                    agency = "Outside agency"
        freq_norm = norm_freq(freq_raw)

        # last_done = the single dated cell in the Calib. On row (cols 4..15)
        last_done = None
        for c in range(4, 16):
            d = xls_date(book, sheet, calib_row, c)
            if d:
                last_done = d
                break

        # Emit one row per dated Due On cell
        nnn += 1
        due_dates = []
        if due_row < sheet.nrows:
            for c in range(4, 16):
                d = xls_date(book, sheet, due_row, c)
                if d:
                    due_dates.append((c, d))

        if not due_dates:
            exceptions.append((CAL_FILE, sheet_name, due_row + 1,
                               f"{clean_name} ({equip_id}) has Calib.On but no dated Due On cell",
                               "No due date present -> not registered"))
        for c, d in due_dates:
            if not in_window(d):
                exceptions.append((CAL_FILE, sheet_name, due_row + 1,
                                   f"{clean_name} Due On {d.isoformat()}",
                                   "Due date outside 2026-01-01..2027-07-31 window -> skipped"))
                continue
            remarks_parts = []
            if freq_raw.strip():
                remarks_parts.append(f"raw freq: {freq_raw.strip()}")
            if agency:
                remarks_parts.append(agency)
            remarks_parts.append(f"sheet: {sheet_name.strip()}")
            rows.append({
                "task_id": f"MIC-{planner_code}-{nnn:03d}-{d.strftime('%Y%m%d')}",
                "department": "Micro",
                "planner_type": planner_type,
                "task_name": ("Calibrate " if activity_type == "calibration" else "Re-validate ") + clean_name,
                "equipment_or_item_id": equip_id,
                "activity_type": activity_type,
                "frequency": freq_norm,
                "due_type": "specific_date",
                "due_date": d.isoformat(),
                "last_done_date": last_done.isoformat() if last_done else "",
                "done_date": "",
                "status": "pending",
                "responsible_email": "MICRO_EMAIL",
                "report_link": "",
                "remarks": "; ".join(remarks_parts),
                "source_file": CAL_FILE,
            })
        r = due_row + 1


parse_matrix("Calibration Planner 26-27 ", "CAL", "Instrument Calibration", "calibration")
parse_matrix("Re-validation Planner 26-27 ", "VAL", "Validation", "re-validation")


# ---------------------------------------------------------------------------
# 2. EM schedules (xlsx, 12 monthly sheets)
# ---------------------------------------------------------------------------
def parse_em(fname, planner_code, planner_type, task_prefix):
    path = os.path.join(ROOT, fname)
    wb = openpyxl.load_workbook(path, data_only=True)
    loc_nnn = {}
    next_nnn = 0
    count_by_month = {}
    for sn in wb.sheetnames:
        ss = wb[sn]
        # find header row + column containing "Date" (some sheets are shifted right)
        header_row = None
        base_col = None
        for r in range(1, ss.max_row + 1):
            for c in range(1, ss.max_column + 1):
                if str(ss.cell(r, c).value).strip().lower() == "date":
                    header_row, base_col = r, c
                    break
            if header_row:
                break
        if header_row is None:
            continue
        month_count = 0
        for r in range(header_row + 1, ss.max_row + 1):
            dcell = ss.cell(r, base_col).value
            fmt = ss.cell(r, base_col + 2).value
            loc = ss.cell(r, base_col + 3).value
            if dcell is None or str(dcell).strip() == "":
                continue
            # skip footer note rows
            if str(dcell).strip().lower().startswith(("note", "ref")):
                continue
            fmt_s = "" if fmt is None else str(fmt).strip()
            loc_s = "" if loc is None else str(loc).strip()
            # non-sampling day: no format AND/OR marked NA
            if not fmt_s or not loc_s or fmt_s.upper() == "NA" or loc_s.upper() == "NA":
                continue
            # parse date
            d = None
            if isinstance(dcell, datetime):
                d = dcell.date()
            elif isinstance(dcell, date):
                d = dcell
            else:
                for fs in ("%d-%b-%Y", "%d-%b-%y", "%d-%B-%Y"):
                    try:
                        d = datetime.strptime(str(dcell).strip(), fs).date()
                        break
                    except ValueError:
                        continue
            if d is None:
                exceptions.append((fname, sn, r, str(dcell), "Unparseable EM date"))
                continue
            if not in_window(d):
                exceptions.append((fname, sn, r, f"{d} {loc_s}", "EM date outside window -> skipped"))
                continue
            if loc_s not in loc_nnn:
                next_nnn += 1
                loc_nnn[loc_s] = next_nnn
            nnn = loc_nnn[loc_s]
            rows.append({
                "task_id": f"MIC-{planner_code}-{nnn:03d}-{d.strftime('%Y%m%d')}",
                "department": "Micro",
                "planner_type": planner_type,
                "task_name": f"{task_prefix} - {loc_s}",
                "equipment_or_item_id": fmt_s,
                "activity_type": "environmental_monitoring",
                "frequency": "per-schedule",
                "due_type": "specific_date",
                "due_date": d.isoformat(),
                "last_done_date": "",
                "done_date": "",
                "status": "pending",
                "responsible_email": "MICRO_EMAIL",
                "report_link": "",
                "remarks": f"HIGH_FREQ; sheet: {sn}",
                "source_file": fname,
            })
            month_count += 1
        count_by_month[sn] = month_count
    return count_by_month


air_counts = parse_em("Micro/EM  Schedule-2026-Air Sampling Monitoring 2026.xlsx",
                      "EMAIR", "EM - Air Sampling", "EM air sampling")
settle_counts = parse_em("Micro/EM  Schedule-2026-Settle Plate Monitoring 2026.xlsx",
                         "EMSET", "EM - Settle Plate", "EM settle plate")


# ---------------------------------------------------------------------------
# Write CSV
# ---------------------------------------------------------------------------
os.makedirs(os.path.dirname(OUT_CSV), exist_ok=True)
with open(OUT_CSV, "w", newline="") as f:
    w = csv.DictWriter(f, fieldnames=COLUMNS, quoting=csv.QUOTE_MINIMAL)
    w.writeheader()
    for row in rows:
        w.writerow(row)

# ---------------------------------------------------------------------------
# Write exceptions / notes
# ---------------------------------------------------------------------------
with open(OUT_EXC, "w") as f:
    f.write("# Micro parsing — exceptions & notes\n\n")
    f.write(f"Generated {datetime.now().isoformat(timespec='seconds')}; reference date 2026-07-15.\n\n")

    f.write("## Missing planners\n\n")
    f.write("- **Media preparation / Growth promotion test (GPT) cycle planner: MISSING.** "
            "The director expected a media-prep / growth-promotion planner from Micro. "
            "None of the three supplied Micro files contains any 'media', 'growth', or 'promotion' "
            "content (searched all sheets of the calibration/validation .xls and both EM .xlsx files). "
            "No rows were fabricated.\n\n")

    f.write("## Assignments to confirm\n\n")
    f.write("- None. All parsed tasks are unambiguously Microbiology-owned "
            "(instrument IDs QCM/QCI, Micro EM schedules). No OWNERSHIP-UNCONFIRMED rows.\n\n")

    f.write("## Exceptions (rows not registered / caveats)\n\n")
    if exceptions:
        f.write("| source_file | sheet | row | cell / content | reason |\n")
        f.write("|---|---|---|---|---|\n")
        for sf, sh, rn, cell, reason in exceptions:
            cell = str(cell).replace("|", "\\|")
            reason = str(reason).replace("|", "\\|")
            f.write(f"| {sf} | {str(sh).strip()} | {rn} | {cell} | {reason} |\n")
    else:
        f.write("None.\n")
    f.write("\n")

# ---------------------------------------------------------------------------
# Console summary
# ---------------------------------------------------------------------------
from collections import Counter
print("TOTAL ROWS:", len(rows))
print("BY planner_type:", dict(Counter(r["planner_type"] for r in rows)))
print("AIR counts by month:", air_counts, "=> total", sum(air_counts.values()))
print("SETTLE counts by month:", settle_counts, "=> total", sum(settle_counts.values()))
print("EXCEPTIONS:", len(exceptions))
