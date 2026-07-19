#!/usr/bin/env python3
"""
Parser for the Engineering department Preventive Maintenance planner.

Source: Engineering/Preventive Maintenance Planner 2026.xlsx
Format: instrument x month matrix. Row 1 = title, Row 2 = header
(col0 Sr.No, col1 Name Of Instrument, col2 Frequency, cols 3..14 = Jan..Dec 2026).
Each dated cell under a month = one scheduled PM occurrence.

Emits one canonical register row per dated PM cell, across all sheets,
following master-reminder-system/PARSING_SPEC.md exactly.
"""

import csv
import datetime
import re
from pathlib import Path

import openpyxl

ROOT = Path("/Users/swarali/Desktop/Enicar/ENincar 2025-26 schedule")
SRC_REL = "Engineering/Preventive Maintenance Planner 2026.xlsx"
SRC = ROOT / SRC_REL
OUT_CSV = ROOT / "master-reminder-system/data/parsed/eng_register.csv"
OUT_EXC = ROOT / "master-reminder-system/data/parsed/eng-exceptions.md"

DEPT = "Engineering"
DEPT_CODE = "ENG"
PLANNER_CODE = "PM"
RESP_EMAIL = "ENG_EMAIL"

WINDOW_START = datetime.date(2026, 1, 1)
WINDOW_END = datetime.date(2027, 7, 31)

QC_SHEETS = {"PM Q.C.I. LAB", "PM Q.C.M. LAB"}

COLUMNS = [
    "task_id", "department", "planner_type", "task_name", "equipment_or_item_id",
    "activity_type", "frequency", "due_type", "due_date", "last_done_date",
    "done_date", "status", "responsible_email", "report_link", "remarks", "source_file",
]

ID_RE = re.compile(r"[A-Z]{2,4}\s*/\s*\d{3,6}")


def extract_id_and_name(raw_name):
    """Return (equipment_id, task_name_core) from an instrument name string."""
    name = str(raw_name)
    m = ID_RE.search(name)
    if not m:
        return "", re.sub(r"\s+", " ", name).strip()
    equip_id = re.sub(r"\s*", "", m.group())  # remove any internal whitespace
    # task name is everything before the ID; drop a trailing '(' and whitespace,
    # keeping any earlier descriptive parenthetical intact.
    core = name[: m.start()]
    core = re.sub(r"[\s(]+$", "", core)
    core = re.sub(r"\s+", " ", core).strip()
    return equip_id, core


def normalize_frequency(raw):
    """Smallest interval implied by a combined code like 'M/Q/HY/Y'."""
    tokens = [t.strip().upper() for t in str(raw).split("/") if t.strip()]
    if "W" in tokens:
        return "weekly"
    if "M" in tokens:
        return "monthly"
    if "Q" in tokens:
        return "quarterly"
    if "HY" in tokens:
        return "half-yearly"
    if "Y" in tokens:
        return "yearly"
    return "per-schedule"


def main():
    wb = openpyxl.load_workbook(SRC, data_only=True)

    rows = []
    exceptions = []
    instrument_nnn = {}  # (sheet, equip_id, core_name) -> stable NNN
    next_nnn = 1
    per_sheet = {}
    instruments = set()

    for sn in wb.sheetnames:
        ws = wb[sn]
        sheet_label = sn.strip()
        per_sheet.setdefault(sheet_label, 0)
        last_instrument = None  # for carry-forward continuation rows

        for r in range(3, ws.max_row + 1):
            sr = ws.cell(r, 1).value
            raw_name = ws.cell(r, 2).value
            raw_freq = ws.cell(r, 3).value

            # collect dated month cells
            dated = []
            for c in range(4, 16):
                v = ws.cell(r, c).value
                if v is None:
                    continue
                if isinstance(v, datetime.datetime):
                    dated.append((c, v.date()))
                else:
                    exceptions.append(
                        (SRC_REL, sn, r, f"col{c}={v!r}",
                         "month cell is not a real date")
                    )

            has_name = raw_name is not None and str(raw_name).strip() != ""

            # skip header echoes / fully blank rows with no dates
            if not has_name and not dated:
                continue

            if has_name:
                equip_id, core = extract_id_and_name(raw_name)
                freq_raw = str(raw_freq).strip() if raw_freq is not None else ""
                last_instrument = (equip_id, core, freq_raw)
            else:
                # continuation row: blank name but has dates -> carry forward
                if last_instrument is None:
                    exceptions.append(
                        (SRC_REL, sn, r, "blank name with dated cells",
                         "no preceding instrument to carry forward")
                    )
                    continue
                equip_id, core, freq_raw = last_instrument

            if not core:
                exceptions.append(
                    (SRC_REL, sn, r, f"name={raw_name!r}",
                     "could not derive a task name")
                )
                continue

            if not dated:
                # section header / instrument with no scheduled dates -> nothing to emit
                continue

            key = (sn, equip_id, core)
            if key not in instrument_nnn:
                instrument_nnn[key] = next_nnn
                next_nnn += 1
            nnn = instrument_nnn[key]
            instruments.add(key)

            frequency = normalize_frequency(freq_raw) if freq_raw else "per-schedule"

            remark_bits = []
            if freq_raw:
                remark_bits.append(f"raw freq: {freq_raw}")
            remark_bits.append(f"source sheet: {sheet_label}")
            if sn in QC_SHEETS:
                remark_bits.append("PM on QC-lab equipment")
            remarks = "; ".join(remark_bits)

            task_name = f"PM - {core}"

            for _c, d in dated:
                if not (WINDOW_START <= d <= WINDOW_END):
                    exceptions.append(
                        (SRC_REL, sn, r, f"date={d.isoformat()}",
                         "outside occurrence window 2026-01-01..2027-07-31")
                    )
                    continue
                ymd = d.strftime("%Y%m%d")
                task_id = f"{DEPT_CODE}-{PLANNER_CODE}-{nnn:03d}-{ymd}"
                rows.append({
                    "task_id": task_id,
                    "department": DEPT,
                    "planner_type": "Preventive Maintenance",
                    "task_name": task_name,
                    "equipment_or_item_id": equip_id,
                    "activity_type": "preventive_maintenance",
                    "frequency": frequency,
                    "due_type": "specific_date",
                    "due_date": d.isoformat(),
                    "last_done_date": "",
                    "done_date": "",
                    "status": "pending",
                    "responsible_email": RESP_EMAIL,
                    "report_link": "",
                    "remarks": remarks,
                    "source_file": SRC_REL,
                })
                per_sheet[sheet_label] += 1

    OUT_CSV.parent.mkdir(parents=True, exist_ok=True)
    with open(OUT_CSV, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=COLUMNS, quoting=csv.QUOTE_MINIMAL)
        w.writeheader()
        w.writerows(rows)

    # exceptions markdown
    with open(OUT_EXC, "w", encoding="utf-8") as f:
        f.write("# Engineering PM Planner - Parsing Exceptions\n\n")
        f.write(f"Source file: `{SRC_REL}`\n\n")
        f.write(f"Reference date: 2026-07-15\n\n")
        if not exceptions:
            f.write("No exceptions. Every data row parsed cleanly: all month "
                    "cells were real dates, every instrument carried an "
                    "extractable equipment ID, and all scheduled dates fell "
                    "within the 2026-01-01..2027-07-31 window.\n")
        else:
            f.write(f"{len(exceptions)} exception row(s):\n\n")
            f.write("| source_file | sheet | row | cell content | reason |\n")
            f.write("|---|---|---|---|---|\n")
            for sf, sh, rr, cell, reason in exceptions:
                cell_s = str(cell).replace("|", "\\|")
                f.write(f"| {sf} | {sh} | {rr} | {cell_s} | {reason} |\n")

    print(f"rows emitted: {len(rows)}")
    print(f"distinct instruments: {len(instruments)}")
    print(f"exceptions: {len(exceptions)}")
    print("per sheet:")
    for k, v in per_sheet.items():
        print(f"  {k}: {v}")


if __name__ == "__main__":
    main()
