#!/usr/bin/env python3
"""
QA department planner parser -> canonical task register CSV.

Sources:
  1. QA/Year 2025-2026 Validation-calibration schedule-QA.xlsx  (sheet '2025-2026')
  2. QA/VMP 2026  (Annexure 1 to 7)-QA.xlsx  (sheets 'Annaexure 1'..'Annaexure 7')

Outputs (paths relative to project root):
  master-reminder-system/data/parsed/qa_register.csv
  master-reminder-system/data/parsed/qa-exceptions.md

Follows master-reminder-system/PARSING_SPEC.md.  Reference date 2026-07-15.
Never guesses: unparseable / ambiguous rows -> exceptions, not register.
"""

import os
import re
import csv
import datetime as dt
import openpyxl

ROOT = "/Users/swarali/Desktop/Enicar/ENincar 2025-26 schedule"
F1_REL = "QA/Year 2025-2026 Validation-calibration schedule-QA.xlsx"
F2_REL = "QA/VMP 2026  (Annexure 1 to 7)-QA.xlsx"
OUT_CSV = os.path.join(ROOT, "master-reminder-system/data/parsed/qa_register.csv")
OUT_EXC = os.path.join(ROOT, "master-reminder-system/data/parsed/qa-exceptions.md")

WINDOW_START = dt.date(2026, 1, 1)
WINDOW_END = dt.date(2027, 7, 31)

COLUMNS = [
    "task_id", "department", "planner_type", "task_name", "equipment_or_item_id",
    "activity_type", "frequency", "due_type", "due_date", "last_done_date",
    "done_date", "status", "responsible_email", "report_link", "remarks",
    "source_file",
]

ID_PREFIXES = ["QCI", "QCM", "STE", "PRE", "UTL", "ULT", "DTH", "DLT", "DLT",
               "GTM", "GPS", "DL"]
# regex to grab an id run in a free-text name (allows ranges like "DTH-01 to DTH -16")
_ID_TOKEN = re.compile(
    r"(?:%s)[\s/–\-]*[A-Za-z]?\d[\w\s/–\-]*?(?=\)|$|,|\s{2,}|[a-z]{4,})"
    % "|".join(sorted(set(ID_PREFIXES), key=len, reverse=True))
)
_ID_ANY = re.compile(
    r"(?:%s)[\s/–\-]*[A-Za-z]?\d[\w/–\-]*"
    % "|".join(sorted(set(ID_PREFIXES), key=len, reverse=True))
)
_DATE_RE = re.compile(r"^\s*(\d{1,2})\.(\d{1,2})\.(\d{2,4})\s*$")

exceptions = []  # list of dicts: source, sheet, rows, cell, reason


def norm(s):
    return re.sub(r"\s+", " ", str(s).strip()) if s is not None else ""


def parse_ddmmyyyy(s):
    """Return (date, ok, reason). ok False -> ambiguous/typo (year not 4-digit or out of range)."""
    m = _DATE_RE.match(str(s))
    if not m:
        return None, False, "unrecognized date format"
    d, mth, y = int(m.group(1)), int(m.group(2)), m.group(3)
    if len(y) != 4:
        return None, False, "year is not 4 digits (likely typo, e.g. '206' -> '2026')"
    y = int(y)
    try:
        return dt.date(y, mth, d), True, ""
    except ValueError as e:
        return None, False, "invalid calendar date (%s)" % e


def extract_id(name):
    """Extract an equipment id run from a free-text instrument name. '' if none."""
    if not name:
        return ""
    m = _ID_TOKEN.search(name)
    if m:
        idtxt = re.sub(r"\s+", " ", m.group(0)).strip(" -/")
        return idtxt
    return ""


# ---------------------------------------------------------------------------
# FILE 1 : Validation-calibration schedule (near-4-column schema)
# ---------------------------------------------------------------------------
def parse_file1():
    rows = []
    wb = openpyxl.load_workbook(os.path.join(ROOT, F1_REL), data_only=True)
    ws = wb["2025-2026"]
    planner = "Validation & Calibration Schedule"
    act_map = {
        "calibration": "calibration",
        "temp. mapping": "temperature_mapping",
        "temp.mapping": "temperature_mapping",
        "validation": "validation",
        "na": "calibration",
    }
    for r in range(2, ws.max_row + 1):
        sr = ws.cell(r, 1).value
        name = norm(ws.cell(r, 2).value)
        perform = ws.cell(r, 3).value
        due = ws.cell(r, 4).value
        rtype = norm(ws.cell(r, 5).value).lower()
        party = norm(ws.cell(r, 6).value)
        if sr is None and not name:
            continue

        # activity type
        act = act_map.get(rtype)
        act_remark = ""
        if act is None:
            # unknown report type
            exceptions.append(dict(source=F1_REL, sheet="2025-2026", rows=str(r),
                                   cell="Types of Reports=%r" % rtype,
                                   reason="unrecognized report type; cannot map activity_type"))
            continue
        if rtype == "na":
            act_remark = "report type 'NA' in sheet -> defaulted to calibration"

        # due date
        if isinstance(due, dt.datetime):
            due_date = due.date()
        elif isinstance(due, dt.date):
            due_date = due
        else:
            dd, ok, reason = parse_ddmmyyyy(due)
            if not ok:
                # suggest reading if it looks like a year typo
                suggest = ""
                mt = re.match(r"^\s*(\d{1,2})\.(\d{1,2})\.(\d{3})\s*$", str(due))
                if mt and mt.group(3).startswith("20") is False and len(mt.group(3)) == 3:
                    suggest = " suggested reading: %s.%s.2026" % (mt.group(1), mt.group(2))
                exceptions.append(dict(source=F1_REL, sheet="2025-2026", rows=str(r),
                                       cell="Due Date=%r (instrument=%r)" % (due, name),
                                       reason="unparseable due date: %s.%s" % (reason, suggest)))
                continue
            due_date = dd

        # last done (perform) date
        last_done = ""
        if isinstance(perform, dt.datetime):
            last_done = perform.date().isoformat()
        elif isinstance(perform, dt.date):
            last_done = perform.isoformat()
        else:
            pd, ok, _ = parse_ddmmyyyy(perform)
            if ok:
                last_done = pd.isoformat()
            # perform typos are non-fatal; leave blank (last_done optional)

        # window filter (yearly items -> only occurrences in window)
        if not (WINDOW_START <= due_date <= WINDOW_END):
            exceptions.append(dict(source=F1_REL, sheet="2025-2026", rows=str(r),
                                   cell="Due Date=%s (instrument=%r)" % (due_date, name),
                                   reason="due date outside occurrence window 2026-01-01..2027-07-31"))
            continue

        eid = extract_id(name)
        remark_bits = ["source sheet '2025-2026'",
                       "frequency inferred yearly from annual calibration schedule"]
        if party and party.lower() != "na":
            remark_bits.append("Party: %s" % party)
        if act_remark:
            remark_bits.append(act_remark)
        remarks = "; ".join(remark_bits)

        nnn = "%03d" % int(sr)
        task_id = "QA-SCHED-%s-%s" % (nnn, due_date.strftime("%Y%m%d"))
        rows.append({
            "task_id": task_id, "department": "QA", "planner_type": planner,
            "task_name": name, "equipment_or_item_id": eid, "activity_type": act,
            "frequency": "yearly", "due_type": "specific_date",
            "due_date": due_date.isoformat(), "last_done_date": last_done,
            "done_date": "", "status": "pending", "responsible_email": "QA_EMAIL",
            "report_link": "", "remarks": remarks, "source_file": F1_REL,
        })
    return rows


# ---------------------------------------------------------------------------
# FILE 2 : VMP annexures (legacy / merged / irregular)
# ---------------------------------------------------------------------------
ACT_LABELS = {
    "re-qualification": "re-qualification",
    "requalification": "re-qualification",
    "re qualification": "re-qualification",
    "validation": "validation",
    "calibration": "calibration",
    "verification": "verification",
    "qualification": "qualification",
}
FREQ_LABELS = ["due date", "done date"]


def classify_freq(text):
    t = re.sub(r"\s+", " ", text.lower())
    if "five" in t:
        return "five-yearly"
    if "twice" in t or "six month" in t or "six-month" in t or "sixmonth" in t:
        return "half-yearly"
    if "two year" in t:
        return "two-yearly"
    if "year" in t:  # 'once in a year', or split 'once in a' + 'year'
        return "yearly"
    if "quarter" in t:
        return "quarterly"
    if "month" in t:
        return "monthly"
    return None


def is_label_cell(v):
    t = norm(v).lower()
    return t in ("due date", "done date")


def cell_is_activity(v):
    t = norm(v).lower()
    return ACT_LABELS.get(t)


def cell_is_freq(v):
    t = norm(v).lower()
    if "date" in t:
        return None
    if any(k in t for k in ("once", "twice", "six month", "six-month",
                            "monthly", "quarter")) or t in ("year", "years"):
        return norm(v)
    return None


def cell_id_tokens(v):
    """Return id fragment string if the cell looks like an equipment id fragment, else None.
    Handles prefixed ids ('PRE 01099'), bare prefixes ('PRE '), bare int fragments (1175),
    and the connector word 'to'."""
    if v is None:
        return None
    if isinstance(v, int):
        return str(v)
    s = norm(v)
    if not s:
        return None
    low = s.lower()
    if low == "to":
        return "to"
    up = s.upper()
    # full prefixed id or bare prefix
    for p in sorted(set(ID_PREFIXES), key=len, reverse=True):
        if up.startswith(p):
            rest = up[len(p):].strip(" /-")
            # accept if there are digits, OR it is a bare prefix (rest empty)
            if rest == "" or re.search(r"\d", up):
                return re.sub(r"\s+", " ", s).strip()
    return None


HEADER_EXACT = {
    "name of equipment/utility/", "name of equipment", "name of instrument",
    "id.no.", "id.no", "id no", "id no.", "qualification/", "qualification/ validation/",
    "frequency", "month", "remark", "remarks", "instrument", "validation/",
    "calibration/", "sr.", "sr", "no", "sr no", "sr no.", "no.",
}
MONTHS = {"jan", "feb", "mar", "apr", "may", "jun",
          "jul", "aug", "sep", "oct", "nov", "dec"}


def parse_vmp():
    rows = []
    wb = openpyxl.load_workbook(os.path.join(ROOT, F2_REL), data_only=True)
    for idx, sn in enumerate(wb.sheetnames, start=1):
        ws = wb[sn]
        planner = "VMP Annexure %d" % idx
        maxc = ws.max_column

        def is_header_row(r):
            hits = 0
            monthhits = 0
            for c in range(2, maxc + 1):
                t = norm(ws.cell(r, c).value).lower()
                if not t:
                    continue
                if t in HEADER_EXACT:
                    hits += 1
                if t in MONTHS:
                    monthhits += 1
            return hits >= 2 or monthhits >= 3

        # locate Due Date and Done Date label rows (row, col)
        due_labels = []
        done_labels = []
        for r in range(1, ws.max_row + 1):
            for c in range(2, maxc + 1):
                t = norm(ws.cell(r, c).value).lower()
                if t == "due date":
                    due_labels.append((r, c))
                elif t == "done date":
                    done_labels.append((r, c))
        done_rows = sorted(set(r for r, _ in done_labels))

        seq = 0
        prev_block_end = 0
        for (dr, dc) in due_labels:
            seq += 1
            # block = rows from prev_block_end+1 .. this item's Done Date row (or dr+3)
            done_after = [r for r in done_rows if r >= dr]
            block_end = done_after[0] if done_after else dr + 3
            block_start = prev_block_end + 1
            prev_block_end = block_end

            # Does the block contain a real item row (bearing an id token or an
            # activity label)?  If so, a no-meat name row is a category header to
            # skip; if NOT (e.g. Annexure 5 'Medical Checkup': no id, no activity
            # label), the lone name row IS the item and must be kept.
            block_has_item = False
            first_item_row = None
            for r in range(block_start, block_end + 1):
                if is_header_row(r):
                    continue
                for c in range(3, maxc + 1):
                    v = ws.cell(r, c).value
                    if cell_is_activity(v) or cell_id_tokens(v) is not None:
                        block_has_item = True
                        first_item_row = r
                        break
                if block_has_item:
                    break

            # gather cells in block
            names = []
            id_frags = []
            act = None
            freq_frags = []  # combine split frequency phrases across cells
            date_cells = []  # (row, col, value)
            for r in range(block_start, block_end + 1):
                if is_header_row(r):
                    continue  # skip sheet column-title / month-band header rows
                # Rows strictly above the item's own id/activity row are category
                # banners for NAME purposes (e.g. 'Electronic Balance', 'Premixing
                # Vessel'); their name text is dropped, but any frequency / date
                # that happens to sit above the id row is still collected.
                pre_item = (first_item_row is not None and r < first_item_row)
                row_has_meat = False
                row_cells = []
                # start at column C (3): column B holds only Sr numbers / letter markers
                for c in range(3, maxc + 1):
                    v = ws.cell(r, c).value
                    if v is None or norm(v) == "":
                        continue
                    row_cells.append((c, v))
                # first pass over row to know if it carries id/act/freq/date
                for c, v in row_cells:
                    if cell_is_activity(v) or cell_is_freq(v) or cell_id_tokens(v):
                        row_has_meat = True
                    if isinstance(v, (dt.datetime, dt.date)):
                        row_has_meat = True
                    elif isinstance(v, str) and _DATE_RE.match(v):
                        row_has_meat = True
                for c, v in row_cells:
                    if is_label_cell(v):
                        continue
                    if norm(v) in ("√", "*"):
                        continue
                    a = cell_is_activity(v)
                    if a:
                        if act is None:
                            act = a
                        continue
                    fr = cell_is_freq(v)
                    if fr:
                        freq_frags.append(fr)
                        continue
                    # date?
                    if isinstance(v, (dt.datetime, dt.date)):
                        dd = v.date() if isinstance(v, dt.datetime) else v
                        date_cells.append((r, c, dd, str(v)))
                        continue
                    if isinstance(v, str) and _DATE_RE.match(v):
                        pdt, ok, _ = parse_ddmmyyyy(v)
                        date_cells.append((r, c, pdt if ok else None, v))
                        continue
                    idt = cell_id_tokens(v)
                    if idt is not None:
                        id_frags.append((r, c, idt))
                        continue
                    # otherwise treat as name text (skip pure ints already handled)
                    if isinstance(v, str):
                        if pre_item:
                            continue  # category-banner text above the item row
                        nv = norm(v)
                        if re.fullmatch(r"-{2,}", nv):
                            continue  # placeholder '--' / '-----' -> not a name
                        names.append((r, c, nv))

            # assemble
            task_name = " ".join(t for _, _, t in sorted(names)).strip()
            task_name = re.sub(r"\s+", " ", task_name)
            id_str = " ".join(t for _, _, t in sorted(id_frags)).strip()
            id_str = re.sub(r"\s+", " ", id_str)
            freq_raw = " ".join(freq_frags).strip()
            freq = classify_freq(freq_raw) if freq_raw else None
            activity = act if act else ""

            rowrange = "%d-%d" % (block_start, block_end)

            # validation: need name + at least one date
            if not date_cells:
                exceptions.append(dict(source=F2_REL, sheet=sn, rows=rowrange,
                                       cell="Due Date label at row %d (name=%r id=%r)" % (dr, task_name, id_str),
                                       reason="no due date value found near Due Date label"))
                continue
            if freq is None:
                exceptions.append(dict(source=F2_REL, sheet=sn, rows=rowrange,
                                       cell="freq=%r name=%r" % (freq_raw, task_name),
                                       reason="frequency phrase missing/unmapped"))
                continue
            if not task_name:
                exceptions.append(dict(source=F2_REL, sheet=sn, rows=rowrange,
                                       cell="Due Date row %d id=%r" % (dr, id_str),
                                       reason="could not resolve task_name"))
                continue

            # any unparseable date in this item -> exception (do not fabricate)
            bad = [(r, c, raw) for (r, c, pdt, raw) in date_cells if pdt is None]
            if bad:
                exceptions.append(dict(source=F2_REL, sheet=sn, rows=rowrange,
                                       cell="; ".join("%s@r%d" % (raw, r) for r, c, raw in bad),
                                       reason="unparseable due date value(s) for item %r" % task_name))
                continue

            # emit one row per due date (occurrence)
            base_nnn = "%03d" % seq
            emitted = 0
            for (r, c, pdt, raw) in sorted(date_cells):
                # window logic
                if freq in ("five-yearly", "two-yearly"):
                    include = True  # next occurrence, even beyond window
                else:
                    include = WINDOW_START <= pdt <= WINDOW_END
                if not include:
                    exceptions.append(dict(source=F2_REL, sheet=sn, rows=rowrange,
                                           cell="due=%s name=%r" % (pdt, task_name),
                                           reason="%s occurrence outside window 2026-01-01..2027-07-31" % freq))
                    continue
                remarks = "%s; raw frequency '%s'" % (sn, freq_raw)
                if not activity:
                    remarks += "; activity label absent in sheet (blank)"
                task_id = "QA-VMP%d-%s-%s" % (idx, base_nnn, pdt.strftime("%Y%m%d"))
                rows.append({
                    "task_id": task_id, "department": "QA", "planner_type": planner,
                    "task_name": task_name, "equipment_or_item_id": id_str,
                    "activity_type": activity, "frequency": freq,
                    "due_type": "specific_date", "due_date": pdt.isoformat(),
                    "last_done_date": "", "done_date": "", "status": "pending",
                    "responsible_email": "QA_EMAIL", "report_link": "",
                    "remarks": remarks, "source_file": F2_REL,
                })
                emitted += 1
    return rows


def main():
    r1 = parse_file1()
    r2 = parse_vmp()
    allrows = r1 + r2

    os.makedirs(os.path.dirname(OUT_CSV), exist_ok=True)
    with open(OUT_CSV, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=COLUMNS, quoting=csv.QUOTE_MINIMAL)
        w.writeheader()
        for row in allrows:
            w.writerow(row)

    # exceptions markdown
    with open(OUT_EXC, "w") as f:
        f.write("# QA Register - Parsing Exceptions & Notes\n\n")
        f.write("Reference date: 2026-07-15.  Generated by qa_parse.py.\n\n")
        f.write("## Exceptions (rows NOT registered)\n\n")
        if not exceptions:
            f.write("_None._\n\n")
        else:
            f.write("| # | Source file | Sheet | Row(s) | Cell / content | Reason |\n")
            f.write("|---|-------------|-------|--------|----------------|--------|\n")
            for i, e in enumerate(exceptions, 1):
                cell = e["cell"].replace("|", "\\|")
                reason = e["reason"].replace("|", "\\|")
                f.write("| %d | %s | %s | %s | %s | %s |\n" %
                        (i, e["source"], e["sheet"], e["rows"], cell, reason))
            f.write("\n")

        f.write("## Missing planners (requested by director, NOT present in QA files)\n\n")
        f.write("The director's QA request referenced a vendor audit planner, APQR "
                "(annual product reviews), and compliance timelines. Neither source file "
                "contains these:\n\n")
        f.write("- **Vendor audit planner** — MISSING. Not found in either the "
                "Validation-calibration schedule or the VMP annexures.\n")
        f.write("- **APQR (Annual Product Reviews)** — MISSING. No product-review schedule "
                "present in either file.\n")
        f.write("- **Compliance timelines** — MISSING. No general compliance-timeline planner "
                "present. (Note: VMP Annexure 5 contains a single 'Medical Checkup' item, "
                "which is a personnel/health item, not the requested compliance-timeline planner.)\n\n")

        f.write("## Assignments to confirm\n\n")
        f.write("- All rows are registered under QA (no cross-department ownership ambiguity "
                "in the source layout).\n")
        f.write("- VMP Annexure 5 'Medical Checkup' (ID '--', no activity label in sheet) IS "
                "registered under QA with a blank activity_type and a remark noting the absent "
                "label (task_id QA-VMP5-001-20270101, due 2027-01-01). Confirm whether this "
                "personnel/health item should instead be owned by HR/Admin, and confirm its "
                "activity_type.\n")
        f.write("- Year-typo dates in exceptions ('.206' endings) are almost certainly '.2026' "
                "but were NOT auto-corrected per the 'never guess' rule; QA should confirm and "
                "re-enter them so the 5 excepted rows can be registered.\n")

    # console summary
    from collections import Counter
    print("FILE1 rows:", len(r1))
    print("VMP rows  :", len(r2))
    print("TOTAL     :", len(allrows))
    print("by planner_type:")
    for k, v in sorted(Counter(r["planner_type"] for r in allrows).items()):
        print("   %-32s %d" % (k, v))
    print("exceptions:", len(exceptions))
    print("by source of exception:")
    for k, v in sorted(Counter(e["source"] for e in exceptions).items()):
        print("   %-55s %d" % (k, v))


if __name__ == "__main__":
    main()
